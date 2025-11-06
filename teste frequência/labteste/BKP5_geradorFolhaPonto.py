BACKUP gerado depois que estabilizou a impressão da GRADE

import fitz  # PyMuPDF
import os
import pandas as pd
from babel.dates import format_date
from datetime import datetime
from openpyxl import load_workbook

# ===============================
# Parâmetros configuráveis
# ===============================
# Posições iniciais para cada turno
x_base_manha = 95
x_base_tarde = 295
x_base_noite = 473

# Posição vertical inicial e espaçamento
y_start = 220
line_height_default = 8       # espaçamento padrão entre linhas
line_height_extra = 10        # espaçamento extra a partir da linha definida
extra_line_start = 4          # aplicar espaçamento extra a partir da 4ª linha (inclusive)
fontsize = 6

# Larguras padrão e extras para colunas
col_width_default = 20
col_width_extra_manha = 30
col_width_extra_tarde = 25
col_width_extra_noite = 25

# Coluna a partir da qual aplicar largura extra (inclusive)
extra_col_start_manha = 2
extra_col_start_tarde = 6
extra_col_start_noite = 3

# Posição vertical para observações (abaixo das grades)
obs_y_offset = 8  # distância abaixo da última linha da grade

# ===============================
# Constantes globais
# ===============================
ANO_REFERENCIA = 2025
MES_REFERENCIA = 10
INICIO_DADOS_PROFESSOR = 10  # Linha onde inicia a lista de professores

def inicializar_programa():
    """
    Inicializa o programa:
    - Verifica arquivo Excel
    - Valida ano/mês
    - Extrai lista de professores e abas
    """
    nome_arquivo_base = f"Base-folhaPonto-{ANO_REFERENCIA}-{MES_REFERENCIA}.xlsx"
    if not os.path.exists(nome_arquivo_base):
        print(f"Arquivo {nome_arquivo_base} não encontrado.")
        return None

    df_parametros = pd.read_excel(nome_arquivo_base, sheet_name='parametros', engine='openpyxl')
    ano_base = int(df_parametros.iloc[3, 1])
    mes_base = int(df_parametros.iloc[4, 1])

    if ano_base != ANO_REFERENCIA or mes_base != MES_REFERENCIA:
        print(f"Erro: Ano/Mês base do arquivo ({ano_base}/{mes_base}) não corresponde ao esperado ({ANO_REFERENCIA}/{MES_REFERENCIA}).")
        return None

    data = datetime(ANO_REFERENCIA, MES_REFERENCIA, 1)
    nome_mes = format_date(data, "MMMM", locale="pt_BR")
    print(f"Atenção, estamos a processar a folha do mês/ano: {nome_mes}/{ANO_REFERENCIA}")

    df_dados = df_parametros.iloc[INICIO_DADOS_PROFESSOR:, [0, 1, 2, 3]]
    df_dados.columns = ['Sequencia', 'Matricula', 'NomeProf', 'Nome da Aba']
    df_dados = df_dados.dropna(how='all')

    dicionario_dados = df_dados.to_dict(orient='records')
    return ANO_REFERENCIA, MES_REFERENCIA, df_parametros, dicionario_dados

def tratar_valor(val):
    """Converte valor para string, substituindo NaN, 'nan' ou vazio por '......'."""
    if pd.isna(val) or str(val).strip().lower() == 'nan' or str(val).strip() == '':
        return '......'
    return str(val).strip()

def extrair_grades(excel_path, aba):
    """
    Extrai as três matrizes (manhã, tarde, noite) da aba do professor.
    Cada matriz é 6x6, sem títulos nem labels, apenas valores.
    """
    wb = load_workbook(excel_path, data_only=True)
    if aba not in wb.sheetnames:
        raise ValueError(f"Aba '{aba}' não encontrada no arquivo Excel.")
    sheet = wb[aba]

    # Função auxiliar para extrair matriz 6x6
    def extrair_matriz(inicio_col, fim_col, inicio_row, fim_row):
        matriz = []
        for r in range(inicio_row, fim_row + 1):
            linha = []
            for c in range(ord(inicio_col), ord(fim_col) + 1):
                cell_value = sheet[f"{chr(c)}{r}"].value
                linha.append(tratar_valor(cell_value))
            matriz.append(linha)
        return matriz

    grade_manha = extrair_matriz('B', 'G', 19, 24)
    grade_tarde = extrair_matriz('H', 'M', 19, 24)
    grade_noite = extrair_matriz('N', 'Q', 19, 24)

    return grade_manha, grade_tarde, grade_noite

def preencher_pdf(dados, modelo_path, saida_path, excel_path):
    nome_abas = dados.get("Nome da Aba")
    if not nome_abas or not isinstance(nome_abas, str):
        print("Nenhuma aba indicada para leitura ou valor inválido.")
        return

    abas_solicitadas = [aba.strip() for aba in nome_abas.split(",") if aba.strip()]
    if not abas_solicitadas:
        print("Nenhuma aba válida encontrada.")
        return

    try:
        abas_disponiveis = pd.ExcelFile(excel_path).sheet_names
    except Exception as e:
        print(f"Erro ao acessar o arquivo Excel: {e}")
        return

    dados_pdf = {
        "NomeProf": tratar_valor(dados.get("NomeProf", "")),
        "Matricula": tratar_valor(dados.get("Matricula", "")),
        "Regime": "",
        "Categoria": "",
        "CargaHoraria": "",
        "HoraAtividade": "",
        "HAE-O": "",
        "HAE-C": "",
        "Disciplinas": [],
        "ObsManha": "",
        "ObsTarde": "",
        "ObsNoite": ""
    }

    grade_manha_final = []
    grade_tarde_final = []
    grade_noite_final = []

    for aba in abas_solicitadas:
        if aba not in abas_disponiveis:
            print(f"Aba '{aba}' não existe no arquivo.")
            continue
        try:
            df_aba = pd.read_excel(excel_path, sheet_name=aba, engine='openpyxl', header=11)
            if df_aba.empty:
                continue
            row = df_aba.iloc[0]

            dados_pdf["Regime"] = tratar_valor(row.get("Regime Juridico", ""))
            dados_pdf["Categoria"] = tratar_valor(row.get("Categoria", ""))
            dados_pdf["CargaHoraria"] = tratar_valor(row.get("Carga Horária", ""))
            dados_pdf["HoraAtividade"] = tratar_valor(row.get("Hora Atividade", ""))
            dados_pdf["HAE-O"] = tratar_valor(row.get("HAE-O", ""))
            dados_pdf["HAE-C"] = tratar_valor(row.get("HAE-C", ""))

            for i in range(1, 7):
                col = f"Disciplina{i}"
                if col in df_aba.columns:
                    val = tratar_valor(row.get(col, ""))
                    dados_pdf["Disciplinas"].append(val)

            # Extrair as grades diretamente da aba
            g_manha, g_tarde, g_noite = extrair_grades(excel_path, aba)
            grade_manha_final = g_manha
            grade_tarde_final = g_tarde
            grade_noite_final = g_noite

            # Extrair observações
            dados_pdf["ObsManha"] = tratar_valor(row.get("Obs-Manha", ""))
            dados_pdf["ObsTarde"] = tratar_valor(row.get("Obs-Tarde", ""))
            dados_pdf["ObsNoite"] = tratar_valor(row.get("Obs-Noite", ""))

        except Exception as e:
            print(f"Erro ao ler aba '{aba}': {e}")

    # Inserir no PDF
    doc = fitz.open(modelo_path)
    page = doc[0]

    # Cabeçalho
    page.insert_text((80, 135), dados_pdf["NomeProf"], fontsize=9)
    page.insert_text((360, 135), dados_pdf["Matricula"], fontsize=8)
    page.insert_text((460, 135), dados_pdf["Regime"], fontsize=8)
    page.insert_text((540, 135), dados_pdf["Categoria"], fontsize=8)

    disciplinas_texto = dados_pdf["Disciplinas"]
    if disciplinas_texto:
        metade = len(disciplinas_texto) // 2 + len(disciplinas_texto) % 2
        linha1 = ", ".join(disciplinas_texto[:metade])
        linha2 = ", ".join(disciplinas_texto[metade:])
        page.insert_text((95, 140), linha1, fontsize=6)
        page.insert_text((95, 146), linha2, fontsize=6)

    page.insert_text((535, 145), dados_pdf["CargaHoraria"], fontsize=9)
    page.insert_text((98, 154), dados_pdf["HoraAtividade"], fontsize=9)
    page.insert_text((295, 154), dados_pdf["HAE-O"], fontsize=9)
    page.insert_text((473, 154), dados_pdf["HAE-C"], fontsize=9)

    # Inserir as três grades lado a lado (sem títulos nem labels)
    for i in range(6):  # 6 linhas
        # Calcular espaçamento vertical com lógica extra
        if i + 1 > extra_line_start:
            y_pos = y_start + (i * line_height_default) + ((i + 1 - extra_line_start) * (line_height_extra - line_height_default))
        else:
            y_pos = y_start + i * line_height_default

        # Manhã
        x_pos = x_base_manha
        for j, val in enumerate(grade_manha_final[i]):
            page.insert_text((x_pos, y_pos), val, fontsize=fontsize)
            if j + 1 == extra_col_start_manha:
                x_pos += col_width_extra_manha
            else:
                x_pos += col_width_default

        # Tarde
        x_pos = x_base_tarde
        for j, val in enumerate(grade_tarde_final[i]):
            page.insert_text((x_pos, y_pos), val, fontsize=fontsize)
            if j + 1 >= extra_col_start_tarde:
                x_pos += col_width_extra_tarde
            else:
                x_pos += col_width_default

        # Noite
        x_pos = x_base_noite
        for j, val in enumerate(grade_noite_final[i]):
            page.insert_text((x_pos, y_pos), val, fontsize=fontsize)
            if j + 1 >= extra_col_start_noite:
                x_pos += col_width_extra_noite
            else:
                x_pos += col_width_default

    # Inserir observações abaixo das grades
    obs_y = y_pos + obs_y_offset
    page.insert_text((55, obs_y), dados_pdf["ObsManha"], fontsize=8)
    page.insert_text((245, obs_y), dados_pdf["ObsTarde"], fontsize=8)
    page.insert_text((435, obs_y), dados_pdf["ObsNoite"], fontsize=8)

    doc.save(saida_path)
    doc.close()

def processamento_central():
    resultado = inicializar_programa()
    if resultado is None:
        return

    ano, mes, df_parametros, dicionario_dados = resultado
    print("Conteúdo extraído da planilha:")
    for item in dicionario_dados:
        print(item)

    pdf_modelo = "_ model.pdf"
    output_dir = "formularios_preenchidos"
    os.makedirs(output_dir, exist_ok=True)
    excel_path = f"Base-folhaPonto-{ano}-{mes}.xlsx"

    for dados in dicionario_dados:
        saida_pdf = os.path.join(output_dir, f"{dados['NomeProf'].replace(' ', '_')}_formulario.pdf")
        preencher_pdf(dados, pdf_modelo, saida_pdf, excel_path)
        print(f"Formulário preenchido salvo em: {saida_pdf}")

def main():
    print("=== Iniciando processamento da folha de ponto ===")
    processamento_central()
    print("=== Fim do processamento ===")

if __name__ == "__main__":
    main()